from openpyxl import load_workbook

def load_workbook_safe(file_path):

    return load_workbook(
        file_path,
        read_only=True,
        data_only=True
    )
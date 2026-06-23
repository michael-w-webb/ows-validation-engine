# test_sheet_inspector.py

from openpyxl import load_workbook

from api.services.workbook_parser.sheet_inspector import (
    extract_workbook_structure
)


# ============================================
# Simple Workbook Parsing
# ============================================

def test_extracts_simple_headers(
    simple_workbook_path
):

    workbook = load_workbook(
        simple_workbook_path,
        read_only=True,
        data_only=True
    )

    result = extract_workbook_structure(
        workbook,
        {
            "Participants": 1
        }
    )

    assert result["Participants"] == [
        "First Name",
        "Last Name",
        "DOB",
        "Zip Code"
    ]


# ============================================
# Offset Header Workbook
# ============================================

def test_offset_headers_require_adjustment(
    offset_headers_path
):

    workbook = load_workbook(
        offset_headers_path,
        read_only=True,
        data_only=True
    )

    incorrect_result = (
        extract_workbook_structure(
            workbook,
            {
                "Participants": 1
            }
        )
    )

    assert incorrect_result[
        "Participants"
    ] != [
        "First Name",
        "Last Name",
        "DOB",
        "Zip"
    ]

    corrected_result = (
        extract_workbook_structure(
            workbook,
            {
                "Participants": 4
            }
        )
    )

    assert corrected_result[
        "Participants"
    ] == [
        "First Name",
        "Last Name",
        "DOB",
        "Zip"
    ]


# ============================================
# Blank Workbook Handling
# ============================================

def test_blank_sheet_returns_empty_structure(
    blank_sheet_workbook_path
):

    workbook = load_workbook(
        blank_sheet_workbook_path,
        read_only=True,
        data_only=True
    )

    result = extract_workbook_structure(
        workbook,
        {
            "Participants": 1
        }
    )

    assert result["Participants"] == []


# ============================================
# Multi-Sheet Workbook
# ============================================

def test_multi_sheet_structure_extraction(
    multi_sheet_workbook_path
):

    workbook = load_workbook(
        multi_sheet_workbook_path,
        read_only=True,
        data_only=True
    )

    result = extract_workbook_structure(
        workbook,
        {
            "Participants": 1,
            "Training": 1,
            "Employment": 1
        }
    )

    assert result["Participants"] == [
        "First Name",
        "Last Name",
        "DOB"
    ]

    assert result["Training"] == [
        "Training Name",
        "Start Date",
        "Completed"
    ]

    assert result["Employment"] == [
        "Employer",
        "Wage",
        "Hire Date"
    ]